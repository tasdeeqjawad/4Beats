import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook

file = '4BeatsQ1.xlsx'
wb = load_workbook(file)

day = datetime.datetime.now().strftime("%A")
ws = wb[day]

driver = webdriver.Chrome(executable_path='/path/to/chromedriver')

for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
    keyword = row[0]
    
    driver.get("http://www.google.com")
    search = driver.find_element_by_name("q")
    search.clear()
    search.send_keys(keyword)
    time.sleep(1)

    search.send_keys(Keys.ARROW_DOWN)
    time.sleep(1)
    suggestions = driver.find_elements_by_css_selector("ul[role='listbox'] li span")

    if suggestions:
        texts = [s.text for s in suggestions]
        long_opt = max(texts, key=len)
        short_opt = min(texts, key=len)
    else:
        long_opt = short_opt = "No suggestions found"

    ws[f'B{row[0].row}'] = long_opt
    ws[f'C{row[0].row}'] = short_opt

wb.save(file)
driver.quit()
